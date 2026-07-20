from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Request, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import Optional, List
from datetime import date, datetime
from zoneinfo import ZoneInfo

_VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

def _now_vn() -> datetime:
    return datetime.now(_VN_TZ).replace(tzinfo=None)
from decimal import Decimal
from pathlib import Path
import io
import secrets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import models, schemas
from database import engine, get_db, Base
from tournament_engine import (
    generate_schedule, generate_group_schedule,
    generate_knockout_from_groups, compute_standings,
)
from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user, require_admin, require_superuser,
    get_club_permission, ClubPermissions,
)

Base.metadata.create_all(bind=engine)

# ── MIGRATION: tự động thêm cột mới vào các bảng cũ khi deploy ──
def _run_migration():
    from sqlalchemy import text

    # Danh sách migration: (table, column, column_def, optional_update_sql)
    migrations = [
        # Bảng users — thêm các cột mới
        ("users", "full_name",    "VARCHAR(100)", None),
        ("users", "is_superuser", "BOOLEAN DEFAULT 0 NOT NULL", None),
        ("users", "member_id",    "INTEGER REFERENCES members(id)", None),
        # Bảng members — thêm rank và club_id
        ("members", "rank",    "VARCHAR(30)", None),
        ("members", "club_id", "INTEGER REFERENCES clubs(id)",
            "UPDATE members SET club_id = (SELECT id FROM clubs LIMIT 1) WHERE club_id IS NULL"),
        # Bảng fee_types — thêm club_id
        ("fee_types", "club_id", "INTEGER REFERENCES clubs(id)",
            "UPDATE fee_types SET club_id = (SELECT id FROM clubs LIMIT 1) WHERE club_id IS NULL"),
        # Bảng transactions — thêm club_id
        ("transactions", "club_id", "INTEGER REFERENCES clubs(id)",
            "UPDATE transactions SET club_id = (SELECT id FROM clubs LIMIT 1) WHERE club_id IS NULL"),
        # Bảng tournaments — thêm club_id
        ("tournaments", "club_id", "INTEGER REFERENCES clubs(id)",
            "UPDATE tournaments SET club_id = (SELECT id FROM clubs LIMIT 1) WHERE club_id IS NULL"),
        # Bảng public_report_tokens — thêm cột slug
        ("public_report_tokens", "slug", "VARCHAR(120)", None),
        # Bảng tournament_participants — thêm hỗ trợ khách mời (player_id)
        ("tournament_participants", "player_id", "INTEGER REFERENCES players(id)", None),
        ("tournament_participants", "partner_player_id", "INTEGER REFERENCES players(id)", None),
        # Bảng club_memberships — thêm telegram_chat_id (đăng nhập bot qua Telegram)
        ("club_memberships", "telegram_chat_id", "INTEGER", None),
    ]

    with engine.connect() as conn:
        for table, column, col_def, update_sql in migrations:
            # Kiểm tra table có tồn tại không
            exists = conn.execute(
                text("SELECT name FROM sqlite_master WHERE type='table' AND name=:t"),
                {"t": table}
            ).fetchone()
            if not exists:
                continue
            # Kiểm tra column đã tồn tại chưa
            cols = [r[1] for r in conn.execute(text(f"PRAGMA table_info({table})")).fetchall()]
            if column not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}"))
                if update_sql:
                    conn.execute(text(update_sql))
        conn.commit()

_run_migration()


def _setup_bot_user():
    """Tự động tạo tài khoản bot nếu BOT_USERNAME + BOT_PASSWORD được set."""
    import os
    bot_username = os.environ.get("BOT_USERNAME")
    bot_password = os.environ.get("BOT_PASSWORD")
    if not bot_username or not bot_password:
        return

    from database import SessionLocal
    db = SessionLocal()
    try:
        # Kiểm tra user đã tồn tại chưa
        existing = db.query(models.User).filter(models.User.username == bot_username).first()
        if not existing:
            bot_user = models.User(
                username=bot_username,
                password_hash=hash_password(bot_password),
                full_name="Telegram Bot",
                role=models.UserRole.admin,
                is_superuser=False,
            )
            db.add(bot_user)
            db.flush()

            # Gán vào tất cả CLB hiện có
            clubs = db.query(models.Club).all()
            for club in clubs:
                membership = models.ClubMembership(
                    user_id=bot_user.id,
                    club_id=club.id,
                    role=models.UserRole.admin,
                    can_view=True, can_create=True, can_edit=True, can_delete=True,
                )
                db.add(membership)

            db.commit()
            print(f"[bot] Đã tạo tài khoản bot '{bot_username}'")
        else:
            # Cập nhật password nếu đổi
            existing.password_hash = hash_password(bot_password)
            # Đảm bảo có membership trong tất cả CLB
            clubs = db.query(models.Club).all()
            for club in clubs:
                ms = db.query(models.ClubMembership).filter(
                    models.ClubMembership.user_id == existing.id,
                    models.ClubMembership.club_id == club.id,
                ).first()
                if not ms:
                    db.add(models.ClubMembership(
                        user_id=existing.id, club_id=club.id,
                        role=models.UserRole.admin,
                        can_view=True, can_create=True, can_edit=True, can_delete=True,
                    ))
            db.commit()
            print(f"[bot] Tài khoản bot '{bot_username}' đã được cập nhật")
    except Exception as e:
        db.rollback()
        print(f"[bot] Lỗi setup bot user: {e}")
    finally:
        db.close()


_setup_bot_user()

app = FastAPI(title="Quản lý CLB Thể thao", version="1.0.0")

# ── Rate limiting (public report endpoints) ──
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = Path(__file__).parent / "static"


# ── AUTH & CLUB ───────────────────────────────────────────

@app.get("/api/club/status")
def club_status(db: Session = Depends(get_db)):
    """Kiểm tra CLB đã được khởi tạo chưa."""
    club = db.query(models.Club).first()
    return {"initialized": club is not None, "club": schemas.ClubOut.from_orm(club) if club else None}


@app.post("/api/club/setup", response_model=schemas.TokenOut)
def club_setup(payload: schemas.ClubSetup, db: Session = Depends(get_db)):
    """Khởi tạo CLB lần đầu + tạo tài khoản admin."""
    if db.query(models.Club).first():
        raise HTTPException(400, "CLB đã được khởi tạo rồi")
    if db.query(models.User).filter(models.User.username == payload.admin_username).first():
        raise HTTPException(400, "Tên đăng nhập đã tồn tại")

    club = models.Club(
        name=payload.club_name, sport=payload.sport,
        description=payload.description, founded_year=payload.founded_year,
        address=payload.address, phone=payload.phone, email=payload.email,
    )
    db.add(club)

    admin = models.User(
        username=payload.admin_username,
        full_name=payload.admin_full_name,
        password_hash=hash_password(payload.admin_password),
        role=models.UserRole.admin,
        is_superuser=True,
    )
    db.add(admin)
    db.commit()
    db.refresh(admin)

    token = create_access_token({"sub": admin.username})
    return {"access_token": token, "user": admin}


@app.post("/api/auth/login", response_model=schemas.TokenOut)
def login(payload: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == payload.username).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(401, "Sai tên đăng nhập hoặc mật khẩu")
    token = create_access_token({"sub": user.username})
    return {"access_token": token, "user": user}


@app.post("/api/auth/register", response_model=schemas.TokenOut)
def register(payload: schemas.RegisterRequest, db: Session = Depends(get_db)):
    if not db.query(models.Club).first():
        raise HTTPException(400, "CLB chưa được khởi tạo")
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(400, "Tên đăng nhập đã tồn tại")
    # Tự đăng ký luôn là role member
    role = models.UserRole.member
    user = models.User(
        username=payload.username,
        full_name=payload.full_name,
        password_hash=hash_password(payload.password),
        role=role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_access_token({"sub": user.username})
    return {"access_token": token, "user": user}


@app.get("/api/auth/me", response_model=schemas.UserOut)
def get_me(current_user: models.User = Depends(get_current_user)):
    return current_user


@app.get("/api/auth/users", response_model=List[schemas.UserOut])
def list_users(db: Session = Depends(get_db),
               current_user: models.User = Depends(require_admin)):
    return db.query(models.User).all()


@app.get("/api/club", response_model=schemas.ClubOut)
def get_club(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    club = db.query(models.Club).first()
    if not club:
        raise HTTPException(404, "Chưa khởi tạo CLB")
    return club


@app.put("/api/club", response_model=schemas.ClubOut)
def update_club(payload: schemas.ClubUpdate, db: Session = Depends(get_db),
                current_user: models.User = Depends(require_admin)):
    club = db.query(models.Club).first()
    if not club:
        raise HTTPException(404, "Chưa khởi tạo CLB")
    for k, v in payload.dict(exclude_none=True).items():
        setattr(club, k, v)
    db.commit()
    db.refresh(club)
    return club


# ── SYSTEM ADMIN ─────────────────────────────────────────

@app.get("/api/admin/users", response_model=List[schemas.UserOut])
def admin_list_users(db: Session = Depends(get_db), su = Depends(require_superuser)):
    return db.query(models.User).order_by(models.User.id).all()

@app.post("/api/admin/users", response_model=schemas.UserOut)
def admin_create_user(payload: schemas.UserCreate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(400, "Tên đăng nhập đã tồn tại")
    role = models.UserRole(payload.role) if payload.role in models.UserRole.__members__ else models.UserRole.member
    user = models.User(
        username=payload.username, full_name=payload.full_name,
        password_hash=hash_password(payload.password),
        role=role, is_superuser=payload.is_superuser,
    )
    db.add(user); db.commit(); db.refresh(user)
    return user

@app.put("/api/admin/users/{uid}", response_model=schemas.UserOut)
def admin_update_user(uid: int, payload: schemas.UserUpdate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    user = db.query(models.User).filter(models.User.id == uid).first()
    if not user: raise HTTPException(404, "Không tìm thấy tài khoản")
    if payload.full_name is not None: user.full_name = payload.full_name
    if payload.role is not None and payload.role in models.UserRole.__members__:
        user.role = models.UserRole(payload.role)
    if payload.is_superuser is not None: user.is_superuser = payload.is_superuser
    if payload.password: user.password_hash = hash_password(payload.password)
    db.commit(); db.refresh(user)
    return user

@app.delete("/api/admin/users/{uid}")
def admin_delete_user(uid: int, db: Session = Depends(get_db), su = Depends(require_superuser)):
    user = db.query(models.User).filter(models.User.id == uid).first()
    if not user: raise HTTPException(404, "Không tìm thấy tài khoản")
    if user.id == su.id: raise HTTPException(400, "Không thể xóa tài khoản đang dùng")
    db.delete(user); db.commit()
    return {"ok": True}

@app.get("/api/admin/clubs", response_model=List[schemas.ClubOut])
def admin_list_clubs(db: Session = Depends(get_db), su = Depends(require_superuser)):
    return db.query(models.Club).order_by(models.Club.id).all()

@app.post("/api/admin/clubs", response_model=schemas.ClubOut)
def admin_create_club(payload: schemas.ClubUpdate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    club = models.Club(**{k: v for k, v in payload.dict().items() if v is not None})
    db.add(club); db.commit(); db.refresh(club)
    return club

@app.put("/api/admin/clubs/{cid}", response_model=schemas.ClubOut)
def admin_update_club(cid: int, payload: schemas.ClubUpdate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    club = db.query(models.Club).filter(models.Club.id == cid).first()
    if not club: raise HTTPException(404, "Không tìm thấy CLB")
    for k, v in payload.dict(exclude_none=True).items():
        setattr(club, k, v)
    db.commit(); db.refresh(club)
    return club

@app.delete("/api/admin/clubs/{cid}")
def admin_delete_club(cid: int, db: Session = Depends(get_db), su = Depends(require_superuser)):
    club = db.query(models.Club).filter(models.Club.id == cid).first()
    if not club: raise HTTPException(404, "Không tìm thấy CLB")
    db.delete(club); db.commit()
    return {"ok": True}

@app.get("/api/admin/memberships", response_model=List[schemas.MembershipOut])
def admin_list_memberships(club_id: Optional[int] = None, db: Session = Depends(get_db), su = Depends(require_superuser)):
    q = db.query(models.ClubMembership)
    if club_id: q = q.filter(models.ClubMembership.club_id == club_id)
    return q.all()

@app.post("/api/admin/memberships", response_model=schemas.MembershipOut)
def admin_create_membership(payload: schemas.MembershipCreate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    existing = db.query(models.ClubMembership).filter(
        models.ClubMembership.user_id == payload.user_id,
        models.ClubMembership.club_id == payload.club_id
    ).first()
    if existing: raise HTTPException(400, "Tài khoản đã được gán cho CLB này")
    role = models.UserRole(payload.role) if payload.role in models.UserRole.__members__ else models.UserRole.member
    m = models.ClubMembership(
        user_id=payload.user_id, club_id=payload.club_id, role=role,
        can_view=payload.can_view, can_create=payload.can_create,
        can_edit=payload.can_edit, can_delete=payload.can_delete,
    )
    db.add(m); db.commit(); db.refresh(m)
    return m

@app.put("/api/admin/memberships/{mid}", response_model=schemas.MembershipOut)
def admin_update_membership(mid: int, payload: schemas.MembershipUpdate, db: Session = Depends(get_db), su = Depends(require_superuser)):
    m = db.query(models.ClubMembership).filter(models.ClubMembership.id == mid).first()
    if not m: raise HTTPException(404, "Không tìm thấy phân quyền")
    if payload.role and payload.role in models.UserRole.__members__:
        m.role = models.UserRole(payload.role)
    for field in ["can_view", "can_create", "can_edit", "can_delete"]:
        val = getattr(payload, field)
        if val is not None: setattr(m, field, val)
    db.commit(); db.refresh(m)
    return m

@app.delete("/api/admin/memberships/{mid}")
def admin_delete_membership(mid: int, db: Session = Depends(get_db), su = Depends(require_superuser)):
    m = db.query(models.ClubMembership).filter(models.ClubMembership.id == mid).first()
    if not m: raise HTTPException(404, "Không tìm thấy phân quyền")
    db.delete(m); db.commit()
    return {"ok": True}


@app.get("/api/my-memberships", response_model=List[schemas.MembershipOut])
def my_memberships(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Tất cả các CLB mà user hiện tại có quyền truy cập."""
    if current_user.is_superuser:
        raise HTTPException(403, "Tài khoản quản trị viên không dùng chế độ thành viên")
    return db.query(models.ClubMembership).filter(
        models.ClubMembership.user_id == current_user.id,
    ).all()


def auto_member_code(db: Session, club_id: int) -> str:
    existing = db.query(models.Member.member_code).filter(models.Member.club_id == club_id).all()
    used = set()
    for (code,) in existing:
        if code and code.startswith("TV") and code[2:].isdigit():
            used.add(int(code[2:]))
    n = 1
    while n in used:
        n += 1
    return f"TV{str(n).zfill(4)}"


# ── MEMBERS ──────────────────────────────────────────────
@app.get("/api/members", response_model=List[schemas.MemberOut])
def list_members(
    search: Optional[str] = None,
    status: Optional[schemas.MemberStatus] = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    q = db.query(models.Member).filter(models.Member.club_id == perms.club_id)
    if search:
        q = q.filter(
            models.Member.full_name.ilike(f"%{search}%") |
            models.Member.member_code.ilike(f"%{search}%") |
            models.Member.phone.ilike(f"%{search}%")
        )
    if status:
        q = q.filter(models.Member.status == status)
    return q.order_by(models.Member.id.desc()).all()


@app.post("/api/members", response_model=schemas.MemberOut, status_code=201)
def create_member(
    data: schemas.MemberCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_create()
    if data.member_code:
        existing = db.query(models.Member).filter(
            models.Member.member_code == data.member_code,
            models.Member.club_id == perms.club_id,
        ).first()
        if existing:
            raise HTTPException(400, "Mã thành viên đã tồn tại trong CLB này")
    member_code = data.member_code or auto_member_code(db, perms.club_id)
    member = models.Member(**data.model_dump(exclude={"member_code"}), member_code=member_code, club_id=perms.club_id)
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


# ── EXCEL TEMPLATE / EXPORT / IMPORT ─────────────────────
# Phải đặt TRƯỚC {member_id} để tránh FastAPI match "template"/"export" như integer
@app.get("/api/members/template")
def download_member_template(perms: ClubPermissions = Depends(get_club_permission)):
    """Tạo và trả về file Excel mẫu để nhập thành viên."""
    perms.require_view()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách thành viên"

    headers = ["Mã thành viên", "Họ và tên (*)", "Ngày sinh", "Số điện thoại",
               "Email", "Ngày vào CLB", "Trạng thái", "Hạng", "Địa chỉ", "Ghi chú"]
    col_widths = [15, 25, 14, 16, 28, 14, 12, 15, 30, 30]
    notes_row = ["VD: HN001", "Nguyễn Văn A", "1990-05-20", "0912345678",
                 "vana@email.com", "2024-01-01", "active", "Chuyên nghiệp",
                 "Hà Nội", "Ghi chú tùy chọn"]

    header_fill = PatternFill("solid", fgColor="1677FF")
    note_fill   = PatternFill("solid", fgColor="E6F0FF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    note_font   = Font(color="595959", name="Arial", size=10, italic=True)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        ws.column_dimensions[cell.column_letter].width = w

    for col, val in enumerate(notes_row, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = note_fill
        cell.font = note_font
        cell.alignment = Alignment(vertical="center")
        cell.border = cell_border

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    ws2 = wb.create_sheet("Hướng dẫn")
    guide = [
        ("HƯỚNG DẪN NHẬP THÀNH VIÊN TỪ EXCEL", None),
        ("", None),
        ("Cột bắt buộc:", None),
        ("  - Họ và tên (*)", "Không được để trống"),
        ("", None),
        ("Cột tùy chọn:", None),
        ("  - Mã thành viên", "Nếu để trống, hệ thống tự tạo"),
        ("  - Ngày sinh / Ngày vào CLB", "Định dạng: YYYY-MM-DD hoặc DD/MM/YYYY"),
        ("  - Trạng thái", "Nhập: active (hoạt động) hoặc inactive (nghỉ). Mặc định: active"),
        ("  - Hạng", "Mới bắt đầu / Nghiệp dư / Bán chuyên / Chuyên nghiệp"),
        ("", None),
        ("Lưu ý:", None),
        ("  - Dòng 1 là tiêu đề, dòng 2 là ví dụ — xóa dòng 2 trước khi nhập thật", None),
        ("  - Nếu Mã thành viên trùng trong CLB, hàng đó sẽ bị bỏ qua", None),
        ("  - Có thể nhập nhiều dòng cùng lúc (tối đa 500 thành viên/lần)", None),
    ]
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 50
    for r, (a, b) in enumerate(guide, start=1):
        ca = ws2.cell(row=r, column=1, value=a)
        if r == 1:
            ca.font = Font(bold=True, size=13, color="1677FF")
        elif a.endswith(":"):
            ca.font = Font(bold=True, size=11)
        if b:
            ws2.cell(row=r, column=2, value=b)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_nhap_thanh_vien.xlsx"},
    )


@app.get("/api/members/export")
def export_members_excel(
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Xuất danh sách thành viên ra file Excel."""
    perms.require_view()
    members = db.query(models.Member).filter(
        models.Member.club_id == perms.club_id
    ).order_by(models.Member.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách thành viên"

    headers = ["Mã thành viên", "Họ và tên", "Ngày sinh", "Số điện thoại",
               "Email", "Ngày vào CLB", "Trạng thái", "Hạng", "Địa chỉ", "Ghi chú"]
    col_widths = [15, 25, 14, 16, 28, 14, 12, 15, 30, 30]

    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_map = {"active": "Hoạt động", "inactive": "Tạm nghỉ", "suspended": "Đình chỉ"}

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    even_fill = PatternFill("solid", fgColor="F5F9FF")
    normal_font = Font(name="Arial", size=10)

    for row_idx, m in enumerate(members, start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        row_data = [
            m.member_code or "",
            m.full_name,
            m.dob.strftime("%d/%m/%Y") if m.dob else "",
            m.phone or "",
            m.email or "",
            m.join_date.strftime("%d/%m/%Y") if m.join_date else "",
            status_map.get(m.status.value if hasattr(m.status, "value") else m.status, m.status),
            m.rank or "",
            m.address or "",
            m.notes or "",
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = normal_font
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date as _date
    filename = f"thanh-vien-{_date.today().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/members/import")
def import_members_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Import thành viên từ file Excel."""
    perms.require_create()

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx hoặc .xls")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ hoặc bị lỗi")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def parse_date(val):
        if not val:
            return None
        if isinstance(val, (date, datetime)):
            return val.date() if isinstance(val, datetime) else val
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def clean(val):
        return str(val).strip() if val is not None else None

    imported, skipped, errors = [], [], []

    for i, row in enumerate(rows, start=2):
        if not any(row):
            continue

        member_code = clean(row[0])
        full_name   = clean(row[1])
        dob         = parse_date(row[2])
        phone       = clean(row[3])
        email       = clean(row[4])
        join_date   = parse_date(row[5])
        status_raw  = clean(row[6]) or "active"
        rank        = clean(row[7])
        address     = clean(row[8])
        notes       = clean(row[9]) if len(row) > 9 else None

        if not full_name:
            errors.append({"row": i, "reason": "Thiếu Họ và tên"})
            continue

        try:
            status = models.MemberStatus(status_raw.lower())
        except ValueError:
            status = models.MemberStatus.active

        if member_code:
            exists = db.query(models.Member).filter(
                models.Member.member_code == member_code,
                models.Member.club_id == perms.club_id,
            ).first()
            if exists:
                skipped.append({"row": i, "name": full_name, "reason": f"Mã '{member_code}' đã tồn tại"})
                continue

        member = models.Member(
            club_id=perms.club_id,
            member_code=member_code,
            full_name=full_name,
            dob=dob,
            phone=phone,
            email=email,
            join_date=join_date,
            status=status,
            rank=rank,
            address=address,
            notes=notes,
        )
        db.add(member)
        imported.append({"row": i, "name": full_name})

    db.commit()

    return {
        "total_rows": len([r for r in rows if any(r)]),
        "imported": len(imported),
        "skipped": len(skipped),
        "errors": len(errors),
        "imported_list": imported,
        "skipped_list": skipped,
        "error_list": errors,
    }


@app.get("/api/members/{member_id}", response_model=schemas.MemberOut)
def get_member(
    member_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    m = db.query(models.Member).filter(
        models.Member.id == member_id,
        models.Member.club_id == perms.club_id,
    ).first()
    if not m:
        raise HTTPException(404, "Không tìm thấy thành viên")
    return m


@app.put("/api/members/{member_id}", response_model=schemas.MemberOut)
def update_member(
    member_id: int,
    data: schemas.MemberUpdate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    m = db.query(models.Member).filter(
        models.Member.id == member_id,
        models.Member.club_id == perms.club_id,
    ).first()
    if not m:
        raise HTTPException(404, "Không tìm thấy thành viên")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(m, k, v)
    db.commit()
    db.refresh(m)
    return m


@app.delete("/api/members/{member_id}", status_code=204)
def delete_member(
    member_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_delete()
    m = db.query(models.Member).filter(
        models.Member.id == member_id,
        models.Member.club_id == perms.club_id,
    ).first()
    if not m:
        raise HTTPException(404, "Không tìm thấy thành viên")
    db.delete(m)
    db.commit()

# ── FEE TYPES ─────────────────────────────────────────────
@app.get("/api/fee-types/template")
def download_fee_type_template(perms: ClubPermissions = Depends(get_club_permission)):
    perms.require_view()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh mục khoản"
    headers = ["Tên khoản (*)", "Loại (*)", "Số tiền mặc định", "Định kỳ", "Mô tả"]
    col_widths = [30, 12, 20, 10, 40]
    notes_row = ["Phí hội viên tháng", "income", "200000", "Có", "Phí đóng hàng tháng"]
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor="1677FF")
    nf = PatternFill("solid", fgColor="E6F0FF")
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = hf; c.border = cb
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    for col, val in enumerate(notes_row, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.fill = nf; c.border = cb
        c.font = Font(color="595959", name="Arial", size=10, italic=True)
    ws.row_dimensions[1].height = 28
    # Sheet hướng dẫn
    ws2 = wb.create_sheet("Hướng dẫn")
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 40
    guide = [
        ("HƯỚNG DẪN NHẬP DANH MỤC KHOẢN", None),
        ("", None),
        ("Cột bắt buộc:", None),
        ("  - Tên khoản (*)", "Không được để trống"),
        ("  - Loại (*)", "Nhập: income (thu) hoặc expense (chi)"),
        ("", None),
        ("Cột tùy chọn:", None),
        ("  - Số tiền mặc định", "Số nguyên, không dấu phẩy. VD: 200000"),
        ("  - Định kỳ", "Nhập: Có hoặc Không (mặc định: Không)"),
        ("  - Mô tả", "Ghi chú thêm về khoản"),
    ]
    for r, (a, b) in enumerate(guide, start=1):
        ca = ws2.cell(row=r, column=1, value=a)
        if r == 1: ca.font = Font(bold=True, size=13, color="1677FF")
        elif a.endswith(":"): ca.font = Font(bold=True, size=11)
        if b: ws2.cell(row=r, column=2, value=b)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_danh_muc_khoan.xlsx"})


@app.get("/api/fee-types/export")
def export_fee_types_excel(db: Session = Depends(get_db), perms: ClubPermissions = Depends(get_club_permission)):
    perms.require_view()
    items = db.query(models.FeeType).filter(models.FeeType.club_id == perms.club_id).order_by(models.FeeType.id).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Danh mục khoản"
    headers = ["Tên khoản", "Loại", "Số tiền mặc định", "Định kỳ", "Mô tả"]
    col_widths = [30, 12, 20, 10, 40]
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor="1677FF"); c.border = cb
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 28
    ef = PatternFill("solid", fgColor="F5F9FF")
    for ri, item in enumerate(items, start=2):
        fill = ef if ri % 2 == 0 else None
        row_data = [item.name, "Thu" if item.type == models.FeeTypeCategory.income else "Chi",
                    float(item.default_amount or 0), "Có" if item.is_recurring else "Không", item.description or ""]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Arial", size=10); c.border = cb
            c.alignment = Alignment(vertical="center")
            if fill: c.fill = fill
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import date as _d
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=danh-muc-khoan-{_d.today()}.xlsx"})


@app.post("/api/fee-types/import")
def import_fee_types_excel(file: UploadFile = File(...), db: Session = Depends(get_db),
                           perms: ClubPermissions = Depends(get_club_permission)):
    perms.require_create()
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx hoặc .xls")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ")
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    def clean(v): return str(v).strip() if v is not None else None
    imported, skipped, errors = [], [], []
    for i, row in enumerate(rows, start=2):
        if not any(row): continue
        name = clean(row[0]); type_raw = clean(row[1]); amount_raw = row[2]
        recurring_raw = clean(row[3]); desc = clean(row[4]) if len(row) > 4 else None
        if not name:
            errors.append({"row": i, "reason": "Thiếu tên khoản"}); continue
        type_raw_lower = (type_raw or "").lower()
        if type_raw_lower in ("income", "thu"): fee_type = models.FeeTypeCategory.income
        elif type_raw_lower in ("expense", "chi"): fee_type = models.FeeTypeCategory.expense
        else:
            errors.append({"row": i, "reason": f"Loại '{type_raw}' không hợp lệ (dùng income/expense hoặc Thu/Chi)"}); continue
        exists = db.query(models.FeeType).filter(
            models.FeeType.name == name, models.FeeType.club_id == perms.club_id).first()
        if exists:
            skipped.append({"row": i, "name": name, "reason": "Tên khoản đã tồn tại"}); continue
        try: amount = float(str(amount_raw).replace(",", "").replace(".", "")) if amount_raw else 0
        except: amount = 0
        is_recurring = str(recurring_raw or "").strip().lower() in ("có", "co", "yes", "true", "1")
        db.add(models.FeeType(club_id=perms.club_id, name=name, type=fee_type,
                               default_amount=amount, is_recurring=is_recurring, description=desc))
        imported.append({"row": i, "name": name})
    db.commit()
    return {"total_rows": len([r for r in rows if any(r)]), "imported": len(imported),
            "skipped": len(skipped), "errors": len(errors),
            "imported_list": imported, "skipped_list": skipped, "error_list": errors}


@app.get("/api/fee-types", response_model=List[schemas.FeeTypeOut])
def list_fee_types(
    type: Optional[schemas.FeeTypeCategory] = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    q = db.query(models.FeeType).filter(models.FeeType.club_id == perms.club_id)
    if type:
        q = q.filter(models.FeeType.type == type)
    return q.order_by(models.FeeType.id.desc()).all()


@app.post("/api/fee-types", response_model=schemas.FeeTypeOut, status_code=201)
def create_fee_type(
    data: schemas.FeeTypeCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_create()
    ft = models.FeeType(**data.model_dump(), club_id=perms.club_id)
    db.add(ft)
    db.commit()
    db.refresh(ft)
    return ft


@app.put("/api/fee-types/{ft_id}", response_model=schemas.FeeTypeOut)
def update_fee_type(
    ft_id: int,
    data: schemas.FeeTypeUpdate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    ft = db.query(models.FeeType).filter(models.FeeType.id == ft_id, models.FeeType.club_id == perms.club_id).first()
    if not ft:
        raise HTTPException(404, "Không tìm thấy loại khoản")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(ft, k, v)
    db.commit()
    db.refresh(ft)
    return ft


@app.delete("/api/fee-types/{ft_id}", status_code=204)
def delete_fee_type(
    ft_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_delete()
    ft = db.query(models.FeeType).filter(models.FeeType.id == ft_id, models.FeeType.club_id == perms.club_id).first()
    if not ft:
        raise HTTPException(404, "Không tìm thấy loại khoản")
    db.delete(ft)
    db.commit()


# ── TRANSACTIONS ──────────────────────────────────────────
@app.get("/api/transactions/template")
def download_transaction_template(db: Session = Depends(get_db), perms: ClubPermissions = Depends(get_club_permission)):
    perms.require_view()
    fee_types = db.query(models.FeeType).filter(models.FeeType.club_id == perms.club_id).all()
    members = db.query(models.Member).filter(models.Member.club_id == perms.club_id).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Giao dịch"
    headers = ["Tên khoản (*)", "Mã thành viên", "Số tiền (*)", "Ngày giao dịch (*)", "Loại (*)", "Phương thức", "Mô tả"]
    col_widths = [30, 16, 15, 18, 12, 18, 40]
    notes_row = ["Phí hội viên tháng", "HN001", "200000",
                 date.today().strftime("%d/%m/%Y"), "income", "Tiền mặt", "Tháng 1/2025"]
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor="1677FF"); c.border = cb
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    for col, val in enumerate(notes_row, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.fill = PatternFill("solid", fgColor="E6F0FF"); c.border = cb
        c.font = Font(color="595959", name="Arial", size=10, italic=True)
    ws.row_dimensions[1].height = 28
    # Sheet danh mục khoản để tham khảo
    ws2 = wb.create_sheet("Danh mục khoản")
    ws2.cell(row=1, column=1, value="Tên khoản").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Loại").font = Font(bold=True)
    ws2.column_dimensions["A"].width = 35; ws2.column_dimensions["B"].width = 12
    for r, ft in enumerate(fee_types, start=2):
        ws2.cell(row=r, column=1, value=ft.name)
        ws2.cell(row=r, column=2, value="income" if ft.type == models.FeeTypeCategory.income else "expense")
    # Sheet danh sách thành viên
    ws3 = wb.create_sheet("Thành viên")
    ws3.cell(row=1, column=1, value="Mã TV").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Họ và tên").font = Font(bold=True)
    ws3.column_dimensions["A"].width = 15; ws3.column_dimensions["B"].width = 30
    for r, m in enumerate(members, start=2):
        ws3.cell(row=r, column=1, value=m.member_code or "")
        ws3.cell(row=r, column=2, value=m.full_name)
    # Sheet hướng dẫn
    ws4 = wb.create_sheet("Hướng dẫn")
    ws4.column_dimensions["A"].width = 45; ws4.column_dimensions["B"].width = 40
    guide = [
        ("HƯỚNG DẪN NHẬP GIAO DỊCH", None),
        ("", None),
        ("Cột bắt buộc:", None),
        ("  - Tên khoản (*)", "Phải khớp đúng tên trong sheet 'Danh mục khoản'"),
        ("  - Số tiền (*)", "Số dương, không dấu phẩy. VD: 200000"),
        ("  - Ngày giao dịch (*)", "Định dạng: DD/MM/YYYY hoặc YYYY-MM-DD"),
        ("  - Loại (*)", "income (thu) hoặc expense (chi)"),
        ("", None),
        ("Cột tùy chọn:", None),
        ("  - Mã thành viên", "Xem danh sách trong sheet 'Thành viên'"),
        ("  - Phương thức", "Tiền mặt / Chuyển khoản / Thẻ (mặc định: Tiền mặt)"),
        ("  - Mô tả", "Ghi chú thêm"),
    ]
    for r, (a, b) in enumerate(guide, start=1):
        ca = ws4.cell(row=r, column=1, value=a)
        if r == 1: ca.font = Font(bold=True, size=13, color="1677FF")
        elif a.endswith(":"): ca.font = Font(bold=True, size=11)
        if b: ws4.cell(row=r, column=2, value=b)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_giao_dich.xlsx"})


@app.get("/api/transactions/export")
def export_transactions_excel(
    month: Optional[int] = None, year: Optional[int] = None,
    db: Session = Depends(get_db), perms: ClubPermissions = Depends(get_club_permission)
):
    perms.require_view()
    q = db.query(models.Transaction).filter(models.Transaction.club_id == perms.club_id)
    if month: q = q.filter(extract("month", models.Transaction.transaction_date) == month)
    if year:  q = q.filter(extract("year",  models.Transaction.transaction_date) == year)
    txs = q.order_by(models.Transaction.transaction_date.desc()).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Giao dịch"
    headers = ["Ngày", "Tên khoản", "Loại", "Thành viên", "Số tiền", "Phương thức", "Mô tả"]
    col_widths = [14, 30, 10, 25, 15, 18, 40]
    thin = Side(style="thin", color="CCCCCC")
    cb = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor="1677FF"); c.border = cb
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 28
    income_fill  = PatternFill("solid", fgColor="F6FFED")
    expense_fill = PatternFill("solid", fgColor="FFF1F0")
    for ri, tx in enumerate(txs, start=2):
        is_income = tx.type == models.FeeTypeCategory.income
        fill = income_fill if is_income else expense_fill
        member_name = tx.member.full_name if tx.member else ""
        row_data = [
            tx.transaction_date.strftime("%d/%m/%Y") if tx.transaction_date else "",
            tx.fee_type.name if tx.fee_type else "",
            "Thu" if is_income else "Chi",
            member_name,
            float(tx.amount or 0),
            tx.payment_method or "Tiền mặt",
            tx.description or "",
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Arial", size=10); c.border = cb
            c.alignment = Alignment(vertical="center"); c.fill = fill
            if col == 5:  # Số tiền — format số
                c.number_format = '#,##0'
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import date as _d
    suffix = f"{year}-{month:02d}" if year and month else str(_d.today())
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=giao-dich-{suffix}.xlsx"})


@app.post("/api/transactions/import")
def import_transactions_excel(file: UploadFile = File(...), db: Session = Depends(get_db),
                               perms: ClubPermissions = Depends(get_club_permission)):
    perms.require_create()
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx hoặc .xls")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ")
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    def clean(v): return str(v).strip() if v is not None else None
    def parse_date(val):
        if not val: return None
        if isinstance(val, (date, datetime)):
            return val.date() if isinstance(val, datetime) else val
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try: return datetime.strptime(s, fmt).date()
            except: pass
        return None
    imported, skipped, errors = [], [], []
    for i, row in enumerate(rows, start=2):
        if not any(row): continue
        fee_name = clean(row[0]); member_code = clean(row[1])
        amount_raw = row[2]; tx_date_raw = row[3]
        type_raw = clean(row[4]); method = clean(row[5]) or "Tiền mặt"
        desc = clean(row[6]) if len(row) > 6 else None
        if not fee_name:
            errors.append({"row": i, "reason": "Thiếu tên khoản"}); continue
        if not amount_raw:
            errors.append({"row": i, "reason": "Thiếu số tiền"}); continue
        tx_date = parse_date(tx_date_raw)
        if not tx_date:
            errors.append({"row": i, "reason": "Ngày giao dịch không hợp lệ"}); continue
        type_lower = (type_raw or "").lower()
        if type_lower in ("income", "thu"): tx_type = models.FeeTypeCategory.income
        elif type_lower in ("expense", "chi"): tx_type = models.FeeTypeCategory.expense
        else:
            errors.append({"row": i, "reason": f"Loại '{type_raw}' không hợp lệ"}); continue
        fee_type = db.query(models.FeeType).filter(
            models.FeeType.name == fee_name, models.FeeType.club_id == perms.club_id).first()
        if not fee_type:
            errors.append({"row": i, "reason": f"Không tìm thấy khoản '{fee_name}'"}); continue
        member_id = None
        if member_code:
            m = db.query(models.Member).filter(
                models.Member.member_code == member_code,
                models.Member.club_id == perms.club_id).first()
            if not m:
                errors.append({"row": i, "reason": f"Không tìm thấy mã TV '{member_code}'"}); continue
            member_id = m.id
        try: amount = float(str(amount_raw).replace(",", ""))
        except: errors.append({"row": i, "reason": "Số tiền không hợp lệ"}); continue
        db.add(models.Transaction(club_id=perms.club_id, fee_type_id=fee_type.id,
            member_id=member_id, type=tx_type, amount=amount,
            transaction_date=tx_date, description=desc, payment_method=method))
        imported.append({"row": i, "name": f"{fee_name} - {amount:,.0f}đ"})
    db.commit()
    return {"total_rows": len([r for r in rows if any(r)]), "imported": len(imported),
            "skipped": len(skipped), "errors": len(errors),
            "imported_list": imported, "skipped_list": skipped, "error_list": errors}


@app.get("/api/transactions", response_model=List[schemas.TransactionOut])
def list_transactions(
    month: Optional[int] = None,
    year: Optional[int] = None,
    type: Optional[schemas.FeeTypeCategory] = None,
    member_id: Optional[int] = None,
    fee_type_id: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    q = db.query(models.Transaction)
    if month:
        q = q.filter(extract("month", models.Transaction.transaction_date) == month)
    if year:
        q = q.filter(extract("year", models.Transaction.transaction_date) == year)
    if type:
        q = q.filter(models.Transaction.type == type)
    if member_id:
        q = q.filter(models.Transaction.member_id == member_id)
    if fee_type_id:
        q = q.filter(models.Transaction.fee_type_id == fee_type_id)
    perms.require_view()
    q = q.filter(models.Transaction.club_id == perms.club_id)
    if search:
        q = q.join(models.Member, models.Transaction.member_id == models.Member.id, isouter=True)\
             .join(models.FeeType, models.Transaction.fee_type_id == models.FeeType.id, isouter=True)\
             .filter(
                models.Transaction.description.ilike(f"%{search}%") |
                models.Member.full_name.ilike(f"%{search}%") |
                models.FeeType.name.ilike(f"%{search}%")
             )
    return q.order_by(models.Transaction.transaction_date.desc()).all()


@app.post("/api/transactions", response_model=schemas.TransactionOut, status_code=201)
def create_transaction(
    data: schemas.TransactionCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_create()
    ft = db.query(models.FeeType).filter(models.FeeType.id == data.fee_type_id, models.FeeType.club_id == perms.club_id).first()
    if not ft:
        raise HTTPException(404, "Loại khoản không tồn tại")
    tx = models.Transaction(**data.model_dump(), type=ft.type, club_id=perms.club_id)
    db.add(tx)
    db.commit()
    db.refresh(tx)
    db.refresh(tx)
    return db.query(models.Transaction).filter(models.Transaction.id == tx.id).first()


@app.put("/api/transactions/{tx_id}", response_model=schemas.TransactionOut)
def update_transaction(
    tx_id: int,
    data: schemas.TransactionCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    tx = db.query(models.Transaction).filter(models.Transaction.id == tx_id, models.Transaction.club_id == perms.club_id).first()
    if not tx:
        raise HTTPException(404, "Không tìm thấy giao dịch")
    ft = db.query(models.FeeType).filter(models.FeeType.id == data.fee_type_id, models.FeeType.club_id == perms.club_id).first()
    if not ft:
        raise HTTPException(404, "Loại khoản không tồn tại")
    for k, v in data.model_dump().items():
        setattr(tx, k, v)
    tx.type = ft.type
    db.commit()
    return db.query(models.Transaction).filter(models.Transaction.id == tx_id).first()


@app.delete("/api/transactions/{tx_id}", status_code=204)
def delete_transaction(
    tx_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_delete()
    tx = db.query(models.Transaction).filter(models.Transaction.id == tx_id, models.Transaction.club_id == perms.club_id).first()
    if not tx:
        raise HTTPException(404, "Không tìm thấy giao dịch")
    db.delete(tx)
    db.commit()


# ── REPORTS ───────────────────────────────────────────────
@app.get("/api/reports/summary")
def report_summary(
    year: int = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    if not year:
        year = _now_vn().year
    results = []
    for month in range(1, 13):
        q = db.query(models.Transaction).filter(
            models.Transaction.club_id == perms.club_id,
            extract("year", models.Transaction.transaction_date) == year,
            extract("month", models.Transaction.transaction_date) == month
        )
        income = sum(float(t.amount) for t in q.filter(models.Transaction.type == "income").all())
        expense = sum(float(t.amount) for t in q.filter(models.Transaction.type == "expense").all())
        results.append({
            "month": month, "year": year,
            "total_income": income,
            "total_expense": expense,
            "balance": income - expense,
        })
    return results


@app.get("/api/reports/overview")
def report_overview(
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    total_members = db.query(func.count(models.Member.id)).filter(models.Member.club_id == perms.club_id).scalar()
    active_members = db.query(func.count(models.Member.id)).filter(models.Member.club_id == perms.club_id, models.Member.status == "active").scalar()
    all_tx = db.query(models.Transaction).filter(models.Transaction.club_id == perms.club_id).all()
    total_income = sum(float(t.amount) for t in all_tx if t.type == "income")
    total_expense = sum(float(t.amount) for t in all_tx if t.type == "expense")
    return {
        "total_members": total_members,
        "active_members": active_members,
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": total_income - total_expense,
    }


@app.get("/api/reports/monthly-detail")
def report_monthly_detail(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Chi tiết thu chi trong một tháng: breakdown theo loại khoản + số dư đầu/cuối kỳ."""
    perms.require_view()

    # Dùng date (không phải datetime) để so sánh với cột Date trong SQLite.
    # SQLite lưu Date dưới dạng chuỗi "YYYY-MM-DD"; nếu so sánh với datetime
    # "YYYY-MM-DD 00:00:00" thì "2026-07-01" < "2026-07-01 00:00:00" = True,
    # khiến giao dịch ngày đầu tháng bị tính nhầm vào kỳ trước.
    from datetime import date as date_type
    start_of_month = date_type(year, month, 1)
    if month == 12:
        end_of_month = date_type(year + 1, 1, 1)
    else:
        end_of_month = date_type(year, month + 1, 1)

    # Tồn quỹ đầu kỳ: tổng đại số toàn bộ giao dịch TRƯỚC tháng này
    # Thu (income) cộng dương, Chi (expense) cộng âm để ra số dư tích lũy
    opening_income = db.query(func.sum(models.Transaction.amount)).filter(
        models.Transaction.club_id == perms.club_id,
        models.Transaction.transaction_date < start_of_month,
        models.Transaction.type == "income",
    ).scalar() or 0

    opening_expense = db.query(func.sum(models.Transaction.amount)).filter(
        models.Transaction.club_id == perms.club_id,
        models.Transaction.transaction_date < start_of_month,
        models.Transaction.type == "expense",
    ).scalar() or 0

    # Tồn quỹ đầu kỳ = tổng thu trước kỳ - tổng chi trước kỳ
    opening_balance = float(opening_income) - float(opening_expense)

    # Giao dịch trong tháng
    txs = db.query(models.Transaction).filter(
        models.Transaction.club_id == perms.club_id,
        models.Transaction.transaction_date >= start_of_month,
        models.Transaction.transaction_date < end_of_month,
    ).all()

    income_by_fee: dict = {}
    expense_by_fee: dict = {}
    for t in txs:
        name = t.fee_type.name if t.fee_type else "Khác"
        bucket = income_by_fee if t.type == "income" else expense_by_fee
        if name not in bucket:
            bucket[name] = {"fee_type": name, "count": 0, "amount": 0}
        bucket[name]["count"] += 1
        bucket[name]["amount"] += float(t.amount)

    total_income = sum(v["amount"] for v in income_by_fee.values())
    total_expense = sum(v["amount"] for v in expense_by_fee.values())

    # Tồn quỹ cuối kỳ = Đầu kỳ + Thu trong kỳ - Chi trong kỳ
    closing_balance = opening_balance + total_income - total_expense

    return {
        "month": month, "year": year,
        "opening_balance": opening_balance,   # Tồn quỹ đầu kỳ (chuyển từ tháng trước)
        "total_income": total_income,
        "total_expense": total_expense,
        "closing_balance": closing_balance,   # Tồn quỹ cuối kỳ (chuyển sang tháng sau)
        "balance": total_income - total_expense,  # Số dư trong kỳ (giữ để backward-compat)
        "transaction_count": len(txs),
        "income_breakdown": sorted(income_by_fee.values(), key=lambda x: -x["amount"]),
        "expense_breakdown": sorted(expense_by_fee.values(), key=lambda x: -x["amount"]),
    }


@app.get("/api/reports/member-contributions")
def member_contributions(
    fee_type_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    q = db.query(
        models.Member.id,
        models.Member.member_code,
        models.Member.full_name,
        models.FeeType.name.label("fee_type_name"),
        func.count(models.Transaction.id).label("transaction_count"),
        func.sum(models.Transaction.amount).label("total_amount"),
    ).join(models.Transaction, models.Transaction.member_id == models.Member.id)\
     .join(models.FeeType, models.FeeType.id == models.Transaction.fee_type_id)\
     .filter(models.Transaction.type == "income", models.Transaction.club_id == perms.club_id)

    if fee_type_id:
        q = q.filter(models.Transaction.fee_type_id == fee_type_id)
    if year:
        q = q.filter(extract("year", models.Transaction.transaction_date) == year)
    if month:
        q = q.filter(extract("month", models.Transaction.transaction_date) == month)

    q = q.group_by(models.Member.id, models.FeeType.id).order_by(models.Member.full_name)
    rows = q.all()
    return [
        {
            "member_id": r.id,
            "member_code": r.member_code,
            "full_name": r.full_name,
            "fee_type_name": r.fee_type_name,
            "transaction_count": r.transaction_count,
            "total_amount": float(r.total_amount or 0),
        }
        for r in rows
    ]


@app.get("/api/reports/fee-status")
def fee_status(
    month: int,
    year: int,
    fee_type_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Danh sách thành viên active đã/chưa đóng một khoản phí trong tháng."""
    perms.require_view()
    members = db.query(models.Member).filter(
        models.Member.club_id == perms.club_id,
        models.Member.status == "active",
    ).order_by(models.Member.full_name).all()
    paid_ids = set(
        r[0] for r in db.query(models.Transaction.member_id).filter(
            models.Transaction.club_id == perms.club_id,
            models.Transaction.fee_type_id == fee_type_id,
            extract("month", models.Transaction.transaction_date) == month,
            extract("year", models.Transaction.transaction_date) == year,
            models.Transaction.member_id.isnot(None),
        ).all()
    )
    result = []
    for m in members:
        result.append({
            "member_id": m.id,
            "member_code": m.member_code,
            "full_name": m.full_name,
            "phone": m.phone,
            "rank": m.rank,
            "paid": m.id in paid_ids,
        })
    paid_count = sum(1 for r in result if r["paid"])
    return {
        "month": month,
        "year": year,
        "total": len(result),
        "paid": paid_count,
        "unpaid": len(result) - paid_count,
        "members": result,
    }


# ── PUBLIC REPORT LINKS ──────────────────────────────────

def _make_slug(club_name: str, token: str) -> str:
    """Tạo slug dạng: ten-clb-viet-lien-khong-dau-XXXXXXXX"""
    import unicodedata, re
    text = unicodedata.normalize("NFD", club_name)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    text = text[:40]  # giới hạn độ dài phần tên
    return f"{text}-{token[:8]}"


# ── PUBLIC REPORT LINKS (CRUD — yêu cầu auth) ────────────

@app.post("/api/report-links")
def create_report_link(
    data: dict,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
    current_user: models.User = Depends(get_current_user),
):
    perms.require_view()
    token_str = secrets.token_urlsafe(32)
    expires_at_val = None
    if data.get("expires_at"):
        try:
            expires_at_val = datetime.fromisoformat(data["expires_at"])
        except Exception:
            expires_at_val = None
    club = db.query(models.Club).filter(models.Club.id == perms.club_id).first()
    slug = _make_slug(club.name if club else "clb", token_str)
    rec = models.PublicReportToken(
        token=token_str,
        slug=slug,
        club_id=perms.club_id,
        label=data.get("label", "Báo cáo công khai"),
        expires_at=expires_at_val,
        created_by=current_user.username,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return {
        "id": rec.id,
        "token": rec.token,
        "slug": rec.slug,
        "label": rec.label,
        "expires_at": rec.expires_at.isoformat() if rec.expires_at else None,
        "is_active": rec.is_active,
        "view_count": rec.view_count,
        "created_at": rec.created_at.isoformat() if rec.created_at else None,
    }


@app.get("/api/report-links")
def list_report_links(
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    recs = db.query(models.PublicReportToken).filter(
        models.PublicReportToken.club_id == perms.club_id
    ).order_by(models.PublicReportToken.created_at.desc()).all()
    return [
        {
            "id": r.id,
            "token": r.token,
            "slug": r.slug,
            "label": r.label,
            "expires_at": r.expires_at.isoformat() if r.expires_at else None,
            "is_active": r.is_active,
            "view_count": r.view_count,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in recs
    ]


@app.patch("/api/report-links/{link_id}/toggle")
def toggle_report_link(
    link_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    rec = db.query(models.PublicReportToken).filter(
        models.PublicReportToken.id == link_id,
        models.PublicReportToken.club_id == perms.club_id,
    ).first()
    if not rec:
        raise HTTPException(404, "Link không tồn tại")
    rec.is_active = not rec.is_active
    db.commit()
    return {"id": rec.id, "is_active": rec.is_active}


@app.delete("/api/report-links/{link_id}")
def delete_report_link(
    link_id: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    rec = db.query(models.PublicReportToken).filter(
        models.PublicReportToken.id == link_id,
        models.PublicReportToken.club_id == perms.club_id,
    ).first()
    if not rec:
        raise HTTPException(404, "Link không tồn tại")
    db.delete(rec)
    db.commit()
    return {"ok": True}


# ── PUBLIC REPORT ENDPOINTS (không cần auth) ─────────────

def _validate_token(slug_or_token: str, db: Session, increment_view: bool = False):
    """Tìm token theo slug (mới) hoặc token đầy đủ (cũ — backward compat)."""
    rec = db.query(models.PublicReportToken).filter(
        models.PublicReportToken.is_active == True,
    ).filter(
        (models.PublicReportToken.slug == slug_or_token) |
        (models.PublicReportToken.token == slug_or_token)
    ).first()
    if not rec:
        raise HTTPException(404, "Link không tồn tại hoặc đã bị vô hiệu hóa")
    if rec.expires_at and rec.expires_at < _now_vn():
        raise HTTPException(410, "Link đã hết hạn")
    if increment_view:
        rec.view_count += 1
        db.commit()
    return rec


@app.get("/api/public/report/{slug}")
@limiter.limit("30/minute")
def public_report_meta(request: Request, slug: str, db: Session = Depends(get_db)):
    rec = _validate_token(slug, db, increment_view=True)
    club = db.query(models.Club).filter(models.Club.id == rec.club_id).first()
    return {
        "label": rec.label,
        "club_name": club.name if club else "",
        "club_sport": club.sport if club else "",
        "expires_at": rec.expires_at.isoformat() if rec.expires_at else None,
        "view_count": rec.view_count,
    }


@app.get("/api/public/report/{slug}/summary")
@limiter.limit("60/minute")
def public_report_summary(request: Request, slug: str, year: int = None, db: Session = Depends(get_db)):
    rec = _validate_token(slug, db)
    if not year:
        year = _now_vn().year
    results = []
    for month in range(1, 13):
        q = db.query(models.Transaction).filter(
            models.Transaction.club_id == rec.club_id,
            extract("year", models.Transaction.transaction_date) == year,
            extract("month", models.Transaction.transaction_date) == month
        )
        income = sum(float(t.amount) for t in q.filter(models.Transaction.type == "income").all())
        expense = sum(float(t.amount) for t in q.filter(models.Transaction.type == "expense").all())
        results.append({
            "month": month, "year": year,
            "total_income": income, "total_expense": expense, "balance": income - expense,
        })
    return results


@app.get("/api/public/report/{slug}/monthly-detail")
@limiter.limit("60/minute")
def public_report_monthly_detail(request: Request, slug: str, month: int, year: int, db: Session = Depends(get_db)):
    rec = _validate_token(slug, db)
    from datetime import date as date_type
    start_of_month = date_type(year, month, 1)
    end_of_month = date_type(year + 1, 1, 1) if month == 12 else date_type(year, month + 1, 1)
    opening_income = db.query(func.sum(models.Transaction.amount)).filter(
        models.Transaction.club_id == rec.club_id,
        models.Transaction.transaction_date < start_of_month,
        models.Transaction.type == "income",
    ).scalar() or 0
    opening_expense = db.query(func.sum(models.Transaction.amount)).filter(
        models.Transaction.club_id == rec.club_id,
        models.Transaction.transaction_date < start_of_month,
        models.Transaction.type == "expense",
    ).scalar() or 0
    opening_balance = float(opening_income) - float(opening_expense)
    txs = db.query(models.Transaction).filter(
        models.Transaction.club_id == rec.club_id,
        models.Transaction.transaction_date >= start_of_month,
        models.Transaction.transaction_date < end_of_month,
    ).all()
    income_by_fee: dict = {}
    expense_by_fee: dict = {}
    for t in txs:
        name = t.fee_type.name if t.fee_type else "Khác"
        bucket = income_by_fee if t.type == "income" else expense_by_fee
        if name not in bucket:
            bucket[name] = {"fee_type": name, "count": 0, "amount": 0}
        bucket[name]["count"] += 1
        bucket[name]["amount"] += float(t.amount)
    total_income = sum(v["amount"] for v in income_by_fee.values())
    total_expense = sum(v["amount"] for v in expense_by_fee.values())
    closing_balance = opening_balance + total_income - total_expense
    return {
        "month": month, "year": year,
        "opening_balance": opening_balance,
        "total_income": total_income, "total_expense": total_expense,
        "closing_balance": closing_balance,
        "balance": total_income - total_expense,
        "transaction_count": len(txs),
        "income_breakdown": sorted(income_by_fee.values(), key=lambda x: -x["amount"]),
        "expense_breakdown": sorted(expense_by_fee.values(), key=lambda x: -x["amount"]),
    }


@app.get("/api/public/report/{slug}/member-contributions")
@limiter.limit("60/minute")
def public_report_member_contributions(
    request: Request, slug: str,
    fee_type_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
):
    rec = _validate_token(slug, db)
    q = db.query(
        models.Member.id,
        models.Member.member_code,
        models.Member.full_name,
        models.FeeType.name.label("fee_type_name"),
        func.count(models.Transaction.id).label("transaction_count"),
        func.sum(models.Transaction.amount).label("total_amount"),
    ).join(models.Transaction, models.Transaction.member_id == models.Member.id)\
     .join(models.FeeType, models.FeeType.id == models.Transaction.fee_type_id)\
     .filter(models.Transaction.type == "income", models.Transaction.club_id == rec.club_id)
    if fee_type_id:
        q = q.filter(models.Transaction.fee_type_id == fee_type_id)
    if year:
        q = q.filter(extract("year", models.Transaction.transaction_date) == year)
    if month:
        q = q.filter(extract("month", models.Transaction.transaction_date) == month)
    rows = q.group_by(models.Member.id, models.FeeType.id).order_by(models.Member.full_name).all()
    return [
        {
            "member_id": r.id, "member_code": r.member_code, "full_name": r.full_name,
            "fee_type_name": r.fee_type_name,
            "transaction_count": r.transaction_count,
            "total_amount": float(r.total_amount or 0),
        }
        for r in rows
    ]


@app.get("/api/public/report/{slug}/fee-status")
@limiter.limit("60/minute")
def public_report_fee_status(
    request: Request, slug: str, month: int, year: int, fee_type_id: int,
    db: Session = Depends(get_db),
):
    rec = _validate_token(slug, db)
    members = db.query(models.Member).filter(
        models.Member.club_id == rec.club_id,
        models.Member.status == "active",
    ).order_by(models.Member.full_name).all()
    paid_ids = set(
        r[0] for r in db.query(models.Transaction.member_id).filter(
            models.Transaction.club_id == rec.club_id,
            models.Transaction.fee_type_id == fee_type_id,
            extract("month", models.Transaction.transaction_date) == month,
            extract("year", models.Transaction.transaction_date) == year,
            models.Transaction.member_id.isnot(None),
        ).all()
    )
    result = [
        {
            "member_id": m.id, "member_code": m.member_code, "full_name": m.full_name,
            "phone": m.phone, "rank": m.rank, "paid": m.id in paid_ids,
        }
        for m in members
    ]
    paid_count = sum(1 for r in result if r["paid"])
    return {
        "month": month, "year": year,
        "total": len(result), "paid": paid_count, "unpaid": len(result) - paid_count,
        "members": result,
    }


@app.get("/api/public/report/{slug}/fee-types")
@limiter.limit("60/minute")
def public_report_fee_types(
    request: Request, slug: str,
    type: Optional[str] = None,
    db: Session = Depends(get_db),
):
    rec = _validate_token(slug, db)
    q = db.query(models.FeeType).filter(models.FeeType.club_id == rec.club_id)
    if type:
        q = q.filter(models.FeeType.type == type)
    fts = q.order_by(models.FeeType.name).all()
    return [{"id": f.id, "name": f.name, "type": f.type} for f in fts]


@app.get("/api/public/report/{slug}/transactions")
@limiter.limit("60/minute")
def public_report_transactions(
    request: Request, slug: str,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    rec = _validate_token(slug, db)
    q = db.query(models.Transaction).filter(models.Transaction.club_id == rec.club_id)
    if month:
        q = q.filter(extract("month", models.Transaction.transaction_date) == month)
    if year:
        q = q.filter(extract("year", models.Transaction.transaction_date) == year)
    txs = q.order_by(models.Transaction.transaction_date.desc()).all()
    return [
        {
            "id": t.id,
            "type": t.type,
            "amount": float(t.amount),
            "transaction_date": t.transaction_date.isoformat() if t.transaction_date else None,
            "description": t.description,
            "payment_method": t.payment_method,
            "fee_type": {"id": t.fee_type.id, "name": t.fee_type.name} if t.fee_type else None,
            "member": {"id": t.member.id, "full_name": t.member.full_name} if t.member else None,
        }
        for t in txs
    ]


@app.get("/api/public/report/{slug}/tournaments")
@limiter.limit("60/minute")
def public_tournaments_list(request: Request, slug: str, db: Session = Depends(get_db)):
    rec = _validate_token(slug, db)
    ts = db.query(models.Tournament).filter(
        models.Tournament.club_id == rec.club_id,
        models.Tournament.status != models.TournamentStatus.draft,
    ).order_by(models.Tournament.id.desc()).all()
    return [
        {
            "id": t.id, "name": t.name, "format": t.format,
            "status": t.status, "team_type": t.team_type,
            "created_at": t.created_at.isoformat() if t.created_at else None,
        }
        for t in ts
    ]


@app.get("/api/public/report/{slug}/tournaments/{tid}", response_model=schemas.TournamentOut)
@limiter.limit("60/minute")
def public_tournament_detail(request: Request, slug: str, tid: int, db: Session = Depends(get_db)):
    rec = _validate_token(slug, db)
    t = db.query(models.Tournament).filter(
        models.Tournament.id == tid,
        models.Tournament.club_id == rec.club_id,
        models.Tournament.status != models.TournamentStatus.draft,
    ).first()
    if not t:
        raise HTTPException(404, "Không tìm thấy giải đấu")
    return t


@app.get("/api/public/report/{slug}/tournaments/{tid}/standings")
@limiter.limit("60/minute")
def public_tournament_standings(
    request: Request, slug: str, tid: int,
    group: Optional[str] = None,
    db: Session = Depends(get_db),
):
    rec = _validate_token(slug, db)
    t = db.query(models.Tournament).filter(
        models.Tournament.id == tid,
        models.Tournament.club_id == rec.club_id,
        models.Tournament.status != models.TournamentStatus.draft,
    ).first()
    if not t:
        raise HTTPException(404, "Không tìm thấy giải đấu")

    matches_q = db.query(models.TournamentMatch).filter(
        models.TournamentMatch.tournament_id == tid,
        models.TournamentMatch.phase == "group",
    )
    if group:
        matches_q = matches_q.filter(models.TournamentMatch.group_name == group)
    matches = matches_q.all()

    participants = t.participants
    if group:
        participants = [p for p in participants if p.group_name == group]

    p_dicts = [{
        "id": p.id, "member_id": p.member_id,
        "full_name": p.member.full_name if p.member else "",
        "team_name": p.team_name or (p.member.full_name if p.member else ""),
        "group_name": p.group_name,
    } for p in participants]

    m_dicts = [{
        "p1_id": m.p1_id, "p2_id": m.p2_id,
        "score1": m.score1, "score2": m.score2, "status": m.status,
    } for m in matches]

    return compute_standings(m_dicts, p_dicts, group=group)


# ── TOURNAMENTS ───────────────────────────────────────────

# ── PLAYERS ───────────────────────────────────────────────

@app.get("/api/players", response_model=List[schemas.PlayerOut])
def list_players(
    type: Optional[str] = None,   # "member" | "guest" | None = all
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    q = db.query(models.Player).filter(models.Player.club_id == perms.club_id)
    if type == "member":
        q = q.filter(models.Player.member_id.isnot(None))
    elif type == "guest":
        q = q.filter(models.Player.member_id.is_(None))
    players = q.order_by(models.Player.name).all()
    return [
        schemas.PlayerOut(
            id=p.id, name=p.name, phone=p.phone, email=p.email,
            rank=p.rank, member_id=p.member_id, club_id=p.club_id,
            is_guest=(p.member_id is None),
        )
        for p in players
    ]


@app.post("/api/players", response_model=schemas.PlayerOut, status_code=201)
def create_player(
    data: schemas.PlayerCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_create()
    # Nếu là thành viên CLB, kiểm tra member thuộc CLB hiện tại
    if data.member_id:
        member = db.query(models.Member).filter(
            models.Member.id == data.member_id,
            models.Member.club_id == perms.club_id,
        ).first()
        if not member:
            raise HTTPException(404, "Thành viên không tồn tại trong CLB")
        # Kiểm tra player record đã tồn tại chưa
        existing = db.query(models.Player).filter(
            models.Player.club_id == perms.club_id,
            models.Player.member_id == data.member_id,
        ).first()
        if existing:
            # Trả về record đã có thay vì báo lỗi
            return schemas.PlayerOut(
                id=existing.id, name=existing.name, phone=existing.phone,
                email=existing.email, rank=existing.rank,
                member_id=existing.member_id, club_id=existing.club_id, is_guest=False,
            )
        name = data.name or member.full_name
    else:
        if not data.name:
            raise HTTPException(422, "Tên là bắt buộc với khách mời")
        name = data.name

    p = models.Player(
        name=name, phone=data.phone, email=data.email,
        rank=data.rank or "Chưa xếp hạng",
        member_id=data.member_id, club_id=perms.club_id,
    )
    db.add(p); db.commit(); db.refresh(p)
    return schemas.PlayerOut(
        id=p.id, name=p.name, phone=p.phone, email=p.email,
        rank=p.rank, member_id=p.member_id, club_id=p.club_id,
        is_guest=(p.member_id is None),
    )


@app.put("/api/players/{pid}", response_model=schemas.PlayerOut)
def update_player(
    pid: int,
    data: schemas.PlayerCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    p = db.query(models.Player).filter(
        models.Player.id == pid, models.Player.club_id == perms.club_id
    ).first()
    if not p:
        raise HTTPException(404, "Không tìm thấy player")
    p.name = data.name or p.name
    p.phone = data.phone
    p.email = data.email
    if data.rank is not None:
        p.rank = data.rank
    db.commit(); db.refresh(p)
    return schemas.PlayerOut(
        id=p.id, name=p.name, phone=p.phone, email=p.email,
        rank=p.rank, member_id=p.member_id, club_id=p.club_id,
        is_guest=(p.member_id is None),
    )


@app.delete("/api/players/{pid}", status_code=204)
def delete_player(
    pid: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_delete()
    p = db.query(models.Player).filter(
        models.Player.id == pid, models.Player.club_id == perms.club_id
    ).first()
    if not p:
        raise HTTPException(404, "Không tìm thấy player")
    # Kiểm tra player đang tham gia giải đấu nào không
    used = db.query(models.TournamentParticipant).filter(
        (models.TournamentParticipant.player_id == pid) |
        (models.TournamentParticipant.partner_player_id == pid)
    ).first()
    if used:
        raise HTTPException(400, "Không thể xóa: player đang tham gia giải đấu")
    db.delete(p); db.commit()


# ── TOURNAMENTS ───────────────────────────────────────────

@app.get("/api/tournaments", response_model=List[schemas.TournamentOut])
def list_tournaments(
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    return db.query(models.Tournament).filter(models.Tournament.club_id == perms.club_id).order_by(models.Tournament.id.desc()).all()


@app.post("/api/tournaments", response_model=schemas.TournamentOut, status_code=201)
def create_tournament(
    data: schemas.TournamentCreate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_create()
    t = models.Tournament(
        club_id=perms.club_id,
        name=data.name, format=data.format,
        team_type=data.team_type,
        pairing_mode=data.pairing_mode, rank_rules=data.rank_rules,
        num_groups=data.num_groups, description=data.description,
    )
    db.add(t); db.flush()

    def _resolve_name(db, member_id=None, player_id=None) -> str:
        """Lấy tên hiển thị từ member hoặc player."""
        if member_id:
            return db.query(models.Member.full_name).filter(models.Member.id == member_id).scalar() or ""
        if player_id:
            return db.query(models.Player.name).filter(models.Player.id == player_id).scalar() or ""
        return ""

    if data.team_type == "doubles" and data.teams:
        # Doubles: mỗi team dict có thể chứa member_id/player_id cho từng vị trí
        for idx, team in enumerate(data.teams):
            m1 = team.get("member_id")
            p1 = team.get("player_id")
            m2 = team.get("partner_member_id")
            p2 = team.get("partner_player_id")
            tname = team.get("team_name") or None
            if not tname:
                n1 = _resolve_name(db, m1, p1)
                n2 = _resolve_name(db, m2, p2)
                tname = f"{n1} / {n2}" if n1 and n2 else (n1 or n2)
            pt = models.TournamentParticipant(
                tournament_id=t.id,
                member_id=m1, player_id=p1,
                partner_member_id=m2, partner_player_id=p2,
                team_name=tname, seed=idx + 1,
            )
            db.add(pt)
    else:
        # Singles: kết hợp member_ids (thành viên) + player_ids (khách mời)
        idx = 0
        for mid in (data.member_ids or []):
            member = db.query(models.Member).filter(models.Member.id == mid).first()
            pt = models.TournamentParticipant(
                tournament_id=t.id, member_id=mid,
                team_name=member.full_name if member else None, seed=idx + 1,
            )
            db.add(pt); idx += 1
        for pid in (data.player_ids or []):
            pname = db.query(models.Player.name).filter(models.Player.id == pid).scalar() or "Khách"
            pt = models.TournamentParticipant(
                tournament_id=t.id, player_id=pid,
                team_name=pname, seed=idx + 1,
            )
            db.add(pt); idx += 1

    db.commit(); db.refresh(t)
    return t


@app.get("/api/tournaments/{tid}", response_model=schemas.TournamentOut)
def get_tournament(
    tid: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid, models.Tournament.club_id == perms.club_id).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")
    return t


@app.put("/api/tournaments/{tid}", response_model=schemas.TournamentOut)
def update_tournament(
    tid: int,
    data: schemas.TournamentUpdate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid, models.Tournament.club_id == perms.club_id).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t


@app.delete("/api/tournaments/{tid}", status_code=204)
def delete_tournament(
    tid: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_delete()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid, models.Tournament.club_id == perms.club_id).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")
    db.delete(t); db.commit()


def _save_matches(db, tid: int, raw_matches: list, offset_idx: int = 0):
    """Lưu list match dicts vào DB, trả về list (match, raw) đã lưu."""
    saved = []
    for m in raw_matches:
        match = models.TournamentMatch(
            tournament_id=tid,
            round_number=m["round_number"], round_name=m.get("round_name"),
            match_number=m["match_number"], phase=m.get("phase", "group"),
            group_name=m.get("group_name"),
            p1_id=m.get("p1_id"), p2_id=m.get("p2_id"),
        )
        db.add(match); db.flush()
        saved.append((match, m))

    for i, (match, m) in enumerate(saved):
        idx = m.get("_next_match_idx")
        slot = m.get("_next_slot")
        if idx is not None:
            real_idx = idx - offset_idx
            if 0 <= real_idx < len(saved):
                match.next_match_id = saved[real_idx][0].id
                match.next_match_slot = slot
    return saved


@app.post("/api/tournaments/{tid}/generate", response_model=schemas.TournamentOut)
def generate_tournament(
    tid: int,
    shuffle: bool = True,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Sinh lịch đấu bảng (xóa lịch cũ nếu có). Với combined: chỉ sinh vòng bảng."""
    perms.require_edit()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")

    # Xóa matches cũ
    db.query(models.TournamentMatch).filter(models.TournamentMatch.tournament_id == tid).delete()

    participants = t.participants
    pid_list = [p.id for p in participants]

    if t.format.value == "combined":
        # Phân bảng: gán group_name rồi sinh round-robin từng bảng
        letters = "ABCDEFGHIJKLMNOP"
        if shuffle:
            import random
            random.shuffle(pid_list)
        for i, p in enumerate(participants):
            p.group_name = letters[i % t.num_groups]
        db.flush()
        # Rebuild pid_list theo thứ tự đã shuffle
        group_map: dict = {}
        for p in participants:
            group_map.setdefault(p.group_name, []).append(p.id)
        raw = generate_group_schedule(pid_list, t.num_groups, shuffle=False)
    else:
        member_ranks = {p.id: (p.member.rank or "") for p in participants}
        raw = generate_schedule(
            format=t.format.value,
            participant_ids=pid_list,
            pairing_mode=t.pairing_mode,
            rank_rules=t.rank_rules,
            member_ranks=member_ranks,
            num_groups=t.num_groups,
            shuffle=shuffle,
        )

    _save_matches(db, tid, raw)
    t.status = models.TournamentStatus.active
    db.commit(); db.refresh(t)
    return t


@app.post("/api/tournaments/{tid}/start-knockout", response_model=schemas.TournamentOut)
def start_knockout(
    tid: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Tính kết quả vòng bảng và sinh lịch knockout (combined format)."""
    perms.require_edit()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")
    if t.format.value != "combined":
        raise HTTPException(400, "Chỉ áp dụng cho thể thức kết hợp")

    # Tính standings từng bảng
    participants = t.participants
    group_names = sorted(set(p.group_name for p in participants if p.group_name))
    group_standings: dict = {}
    for gname in group_names:
        gparticipants = [p for p in participants if p.group_name == gname]
        gmatches = db.query(models.TournamentMatch).filter(
            models.TournamentMatch.tournament_id == tid,
            models.TournamentMatch.phase == "group",
            models.TournamentMatch.group_name == gname,
        ).all()
        p_dicts = [{
            "id": p.id, "member_id": p.member_id,
            "full_name": p.member.full_name if p.member else "",
            "team_name": p.team_name or (p.member.full_name if p.member else ""),
            "group_name": p.group_name,
        } for p in gparticipants]
        m_dicts = [{
            "p1_id": m.p1_id, "p2_id": m.p2_id,
            "score1": m.score1, "score2": m.score2, "status": m.status,
        } for m in gmatches]
        group_standings[gname] = compute_standings(m_dicts, p_dicts, group=None)

    # Số trận đã có
    existing_count = db.query(func.count(models.TournamentMatch.id)).filter(
        models.TournamentMatch.tournament_id == tid
    ).scalar() or 0

    raw_ko = generate_knockout_from_groups(group_standings, existing_count)

    # Điều chỉnh index offset vì _next_match_idx tính từ 0 trong raw_ko
    saved_ko = []
    for m in raw_ko:
        match = models.TournamentMatch(
            tournament_id=tid,
            round_number=m["round_number"], round_name=m.get("round_name"),
            match_number=m["match_number"], phase="knockout",
            group_name=None,
            p1_id=m.get("p1_id"), p2_id=m.get("p2_id"),
        )
        db.add(match); db.flush()
        saved_ko.append((match, m))

    for i, (match, m) in enumerate(saved_ko):
        idx = m.get("_next_match_idx")
        slot = m.get("_next_slot")
        if idx is not None and 0 <= idx < len(saved_ko):
            match.next_match_id = saved_ko[idx][0].id
            match.next_match_slot = slot

    db.commit(); db.refresh(t)
    return t


@app.post("/api/tournaments/{tid}/matches/{mid}/score", response_model=schemas.MatchOut)
def update_score(
    tid: int,
    mid: int,
    data: schemas.ScoreUpdate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    match = db.query(models.TournamentMatch).filter(
        models.TournamentMatch.id == mid,
        models.TournamentMatch.tournament_id == tid
    ).first()
    if not match: raise HTTPException(404, "Không tìm thấy trận đấu")

    match.score1 = data.score1
    match.score2 = data.score2
    match.status = models.MatchStatus.completed

    # Xác định winner
    if data.score1 > data.score2:
        match.winner_id = match.p1_id
    elif data.score2 > data.score1:
        match.winner_id = match.p2_id
    else:
        match.winner_id = None  # hòa — knockout sẽ cần xử lý thủ công

    # Đưa winner vào trận tiếp theo (knockout)
    if match.winner_id and match.next_match_id:
        next_m = db.query(models.TournamentMatch).filter(
            models.TournamentMatch.id == match.next_match_id
        ).first()
        if next_m:
            if match.next_match_slot == 1:
                next_m.p1_id = match.winner_id
            else:
                next_m.p2_id = match.winner_id

    db.commit(); db.refresh(match)
    return match


def _resolve_display_name(db, member_id=None, player_id=None) -> str:
    if member_id:
        return db.query(models.Member.full_name).filter(models.Member.id == member_id).scalar() or ""
    if player_id:
        return db.query(models.Player.name).filter(models.Player.id == player_id).scalar() or ""
    return ""


@app.patch("/api/tournaments/{tid}/participants/{pid}", response_model=schemas.ParticipantOut)
def replace_participant_slot(
    tid: int, pid: int,
    data: schemas.ParticipantSlotUpdate,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Thay người chơi ở 1 vị trí (main hoặc partner) trong 1 đội — dùng khi có
    người không thể tiếp tục thi đấu giữa giải. Giữ nguyên lịch sử các trận đã đấu
    (điểm/thắng-thua vẫn gắn với participant_id, chỉ đổi tên người ở vị trí đó)."""
    perms.require_edit()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid, models.Tournament.club_id == perms.club_id).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")
    p = db.query(models.TournamentParticipant).filter(
        models.TournamentParticipant.id == pid, models.TournamentParticipant.tournament_id == tid,
    ).first()
    if not p: raise HTTPException(404, "Không tìm thấy đội/người chơi")
    if data.slot not in ("main", "partner"):
        raise HTTPException(400, "slot phải là 'main' hoặc 'partner'")
    if not data.member_id and not data.player_id:
        raise HTTPException(400, "Cần chọn thành viên hoặc khách mời để thay")
    if data.member_id and data.player_id:
        raise HTTPException(400, "Chỉ chọn 1 trong 2: thành viên hoặc khách mời")

    # Không cho trùng người đã có mặt ở vị trí khác trong cùng giải
    others = db.query(models.TournamentParticipant).filter(
        models.TournamentParticipant.tournament_id == tid,
        models.TournamentParticipant.id != pid,
    ).all()
    for o in others:
        for mid, plid in [(o.member_id, o.player_id), (o.partner_member_id, o.partner_player_id)]:
            if data.member_id and mid == data.member_id:
                raise HTTPException(400, "Thành viên này đã tham gia giải ở một đội khác")
            if data.player_id and plid == data.player_id:
                raise HTTPException(400, "Khách mời này đã tham gia giải ở một đội khác")
    same_team_other = (p.partner_member_id, p.partner_player_id) if data.slot == "main" else (p.member_id, p.player_id)
    if data.member_id and same_team_other[0] == data.member_id:
        raise HTTPException(400, "Không thể chọn trùng người còn lại trong cùng đội")
    if data.player_id and same_team_other[1] == data.player_id:
        raise HTTPException(400, "Không thể chọn trùng người còn lại trong cùng đội")

    if data.slot == "main":
        p.member_id, p.player_id = data.member_id, data.player_id
    else:
        p.partner_member_id, p.partner_player_id = data.member_id, data.player_id

    if data.team_name:
        p.team_name = data.team_name
    else:
        n1 = _resolve_display_name(db, p.member_id, p.player_id)
        n2 = _resolve_display_name(db, p.partner_member_id, p.partner_player_id)
        p.team_name = f"{n1} / {n2}" if n1 and n2 else (n1 or n2)

    db.commit(); db.refresh(p)
    return p


@app.get("/api/tournaments/{tid}/standings")
def get_standings(
    tid: int,
    group: Optional[str] = None,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_view()
    t = db.query(models.Tournament).filter(models.Tournament.id == tid).first()
    if not t: raise HTTPException(404, "Không tìm thấy giải đấu")

    matches_q = db.query(models.TournamentMatch).filter(
        models.TournamentMatch.tournament_id == tid,
        models.TournamentMatch.phase == "group",
    )
    if group:
        matches_q = matches_q.filter(models.TournamentMatch.group_name == group)
    matches = matches_q.all()

    participants = t.participants
    if group:
        participants = [p for p in participants if p.group_name == group]

    p_dicts = [{
        "id": p.id, "member_id": p.member_id,
        "full_name": p.member.full_name if p.member else "",
        "team_name": p.team_name or (p.member.full_name if p.member else ""),
        "group_name": p.group_name,
    } for p in participants]

    m_dicts = [{
        "p1_id": m.p1_id, "p2_id": m.p2_id,
        "score1": m.score1, "score2": m.score2, "status": m.status,
    } for m in matches]

    return compute_standings(m_dicts, p_dicts, group=group)


# ── BOT CONFIG ────────────────────────────────────────────
def _require_superuser_club_id(
    x_club_id: Optional[int] = Header(default=None, alias="X-Club-ID"),
    current_user: models.User = Depends(get_current_user),
) -> int:
    """Chỉ superuser được ghi bot-config; club_id lấy từ header."""
    if not current_user.is_superuser:
        raise HTTPException(403, "Chỉ Quản trị viên hệ thống mới được cấu hình Bot")
    if not x_club_id:
        raise HTTPException(400, "Thiếu header X-Club-ID")
    return x_club_id


@app.get("/api/bot-config")
def get_bot_config(
    x_club_id: Optional[int] = Header(default=None, alias="X-Club-ID"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Đọc bot-config: superuser được đọc mọi CLB; thành viên thường đọc CLB của mình."""
    if not x_club_id:
        raise HTTPException(400, "Thiếu header X-Club-ID")
    if not current_user.is_superuser:
        membership = db.query(models.ClubMembership).filter(
            models.ClubMembership.user_id == current_user.id,
            models.ClubMembership.club_id == x_club_id,
        ).first()
        if not membership:
            raise HTTPException(403, "Không có quyền truy cập CLB này")
    rows = db.query(models.BotConfig).filter(models.BotConfig.club_id == x_club_id).all()
    return {r.key: r.value for r in rows}


@app.put("/api/bot-config")
def put_bot_config(
    body: dict,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    perms.require_edit()
    club_id = perms.club_id
    for key, value in body.items():
        row = db.query(models.BotConfig).filter(
            models.BotConfig.club_id == club_id,
            models.BotConfig.key == key,
        ).first()
        if row:
            row.value = str(value) if value is not None else None
        else:
            db.add(models.BotConfig(club_id=club_id, key=key, value=str(value) if value is not None else None))
    db.commit()
    rows = db.query(models.BotConfig).filter(models.BotConfig.club_id == club_id).all()
    return {r.key: r.value for r in rows}


# ── FEE REMINDER ENDPOINTS ───────────────────────────────

import os as _os
import urllib.request as _urllib_req
import urllib.parse as _urllib_parse
import json as _json
from calendar import monthrange as _monthrange

_INTERNAL_SECRET = _os.environ.get("INTERNAL_SECRET", "")


def _check_internal_secret(x_internal_secret: str = Header(None)):
    if not _INTERNAL_SECRET or x_internal_secret != _INTERNAL_SECRET:
        raise HTTPException(403, "Forbidden")


@app.patch("/api/my-memberships/telegram-chat-id")
def save_telegram_chat_id(
    body: dict,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Bot gọi endpoint này sau khi admin chọn CLB để lưu telegram_chat_id."""
    chat_id = body.get("chat_id")
    if not chat_id:
        raise HTTPException(400, "Thiếu chat_id")
    membership = db.query(models.ClubMembership).filter(
        models.ClubMembership.user_id == perms.membership.user_id,
        models.ClubMembership.club_id == perms.club_id,
    ).first()
    if not membership:
        raise HTTPException(404, "Không tìm thấy membership")
    membership.telegram_chat_id = int(chat_id)
    db.commit()
    return {"ok": True}


def _build_reminder_data(month: int, year: int, db: Session):
    """Trả về danh sách CLB → fee_types có remind_enabled + thành viên chưa đóng + chat_id admin."""
    fee_types = db.query(models.FeeType).filter(
        models.FeeType.remind_enabled == True,
        models.FeeType.type == models.FeeTypeCategory.income,
    ).all()

    result = []
    for ft in fee_types:
        club_id = ft.club_id
        paid_ids = set(
            r[0] for r in db.query(models.Transaction.member_id).filter(
                models.Transaction.club_id == club_id,
                models.Transaction.fee_type_id == ft.id,
                extract("month", models.Transaction.transaction_date) == month,
                extract("year", models.Transaction.transaction_date) == year,
                models.Transaction.member_id.isnot(None),
            ).all()
        )
        unpaid_members = db.query(models.Member).filter(
            models.Member.club_id == club_id,
            models.Member.status == "active",
            models.Member.id.notin_(paid_ids),
        ).order_by(models.Member.full_name).all()

        admin_memberships = db.query(models.ClubMembership).filter(
            models.ClubMembership.club_id == club_id,
            models.ClubMembership.role == models.UserRole.admin,
            models.ClubMembership.telegram_chat_id.isnot(None),
        ).all()
        chat_ids = [m.telegram_chat_id for m in admin_memberships]

        club = db.query(models.Club).filter(models.Club.id == club_id).first()
        result.append({
            "club_id": club_id,
            "club_name": club.name if club else str(club_id),
            "fee_type_id": ft.id,
            "fee_type_name": ft.name,
            "month": month,
            "year": year,
            "unpaid_count": len(unpaid_members),
            "unpaid_members": [{"id": m.id, "full_name": m.full_name, "phone": m.phone} for m in unpaid_members],
            "admin_chat_ids": chat_ids,
        })
    return result


@app.get("/api/fee-reminders/preview")
def preview_fee_reminders_frontend(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Xem trước danh sách thành viên chưa đóng phí — admin CLB."""
    perms.require_view()
    data = _build_reminder_data(month, year, db)
    return [item for item in data if item["club_id"] == perms.club_id]


@app.post("/api/fee-reminders/send")
def send_fee_reminders_frontend(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    perms: ClubPermissions = Depends(get_club_permission),
):
    """Gửi nhắc đóng phí thủ công — admin CLB."""
    perms.require_edit()
    bot_token = _os.environ.get("TELEGRAM_BOT_TOKEN") or _os.environ.get("BOT_TOKEN", "")
    if not bot_token:
        raise HTTPException(500, "BOT_TOKEN chưa được cấu hình")
    data = [item for item in _build_reminder_data(month, year, db) if item["club_id"] == perms.club_id]
    sent_count = 0
    errors = []
    for item in data:
        if not item["admin_chat_ids"]:
            continue
        for chat_id in item["admin_chat_ids"]:
            unpaid_count = item["unpaid_count"]
            if unpaid_count > 0:
                names = "\n".join(
                    f"  • {m['full_name']}" + (f" ({m['phone']})" if m.get("phone") else "")
                    for m in item["unpaid_members"]
                )
                text = (
                    f"🔔 *Nhắc đóng phí tháng {month}/{year}*\n"
                    f"CLB: {item['club_name']}\n"
                    f"Khoản: {item['fee_type_name']}\n\n"
                    f"Thành viên chưa đóng ({unpaid_count} người):\n{names}"
                )
            else:
                text = (
                    f"✅ *Phí tháng {month}/{year} — {item['fee_type_name']}*\n"
                    f"CLB: {item['club_name']}\n\n"
                    f"Tất cả thành viên đã đóng phí!"
                )
            try:
                payload = _json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}).encode()
                req = _urllib_req.Request(
                    f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    data=payload,
                    headers={"Content-Type": "application/json"},
                )
                with _urllib_req.urlopen(req, timeout=10) as resp:
                    resp.read()
                sent_count += 1
            except Exception as exc:
                errors.append({"chat_id": chat_id, "error": str(exc)})
    return {"sent": sent_count, "errors": errors}


@app.get("/api/internal/fee-reminders")
def preview_fee_reminders(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    _: None = Depends(_check_internal_secret),
):
    return _build_reminder_data(month, year, db)


@app.post("/api/internal/send-fee-reminder")
def send_fee_reminder(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    _: None = Depends(_check_internal_secret),
):
    bot_token = _os.environ.get("TELEGRAM_BOT_TOKEN") or _os.environ.get("BOT_TOKEN", "")
    if not bot_token:
        raise HTTPException(500, "BOT_TOKEN chưa được cấu hình")

    data = _build_reminder_data(month, year, db)
    today = date.today()
    sent_count = 0
    skipped_count = 0
    errors = []

    for item in data:
        if not item["admin_chat_ids"]:
            continue
        if item["unpaid_count"] == 0:
            continue

        for chat_id in item["admin_chat_ids"]:
            # Kiểm tra đã gửi hôm nay chưa
            existing = db.query(models.ReminderLog).filter(
                models.ReminderLog.club_id == item["club_id"],
                models.ReminderLog.fee_type_id == item["fee_type_id"],
                models.ReminderLog.month == month,
                models.ReminderLog.year == year,
                models.ReminderLog.send_date == today,
            ).first()
            if existing:
                skipped_count += 1
                continue

            names = "\n".join(
                f"  • {m['full_name']}" + (f" ({m['phone']})" if m.get("phone") else "")
                for m in item["unpaid_members"]
            )
            text = (
                f"🔔 *Nhắc đóng phí tháng {month}/{year}*\n"
                f"CLB: {item['club_name']}\n"
                f"Khoản: {item['fee_type_name']}\n\n"
                f"Thành viên chưa đóng ({item['unpaid_count']} người):\n{names}"
            )

            try:
                payload = _json.dumps({
                    "chat_id": chat_id,
                    "text": text,
                    "parse_mode": "Markdown",
                }).encode()
                req = _urllib_req.Request(
                    f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    data=payload,
                    headers={"Content-Type": "application/json"},
                )
                with _urllib_req.urlopen(req, timeout=10) as resp:
                    resp.read()

                # Ghi log (UNIQUE constraint tránh trùng)
                log = models.ReminderLog(
                    club_id=item["club_id"],
                    fee_type_id=item["fee_type_id"],
                    month=month,
                    year=year,
                    send_date=today,
                )
                db.add(log)
                db.commit()
                sent_count += 1
            except Exception as exc:
                errors.append({"chat_id": chat_id, "error": str(exc)})

    return {
        "sent": sent_count,
        "skipped_already_sent_today": skipped_count,
        "errors": errors,
    }


# ── SERVE REACT FRONTEND ──────────────────────────────────
if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=STATIC_DIR / "assets"), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        return FileResponse(STATIC_DIR / "index.html")
